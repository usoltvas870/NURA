"""NURA Video Production Pipeline.

Двухэтапный конвейер:
  Этап 1 (--dry-run): тренды → AI → JSON-сценарии
  Этап 2: JSON-сценарии → сборка видео

Usage:
    # Только генерация сценариев (режим A — полуавтомат)
    python scripts/run_pipeline.py --dry-run

    # Генерация + сборка (если все файлы готовы)
    python scripts/run_pipeline.py

    # Из конкретного файла трендов
    python scripts/run_pipeline.py --from-trend ../nura-trend-radar/data/trend_top.json

    # Только топ-3
    python scripts/run_pipeline.py --top 3

    # Валидация сгенерированных сценариев после dry-run
    python scripts/run_pipeline.py --validate
"""
import argparse
import asyncio
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "nura_app"))

from core.services.video_pipeline import VideoPipeline, NURA_ROOT, JOBS_DIR


async def main():
    parser = argparse.ArgumentParser(description="NURA Video Production Pipeline")
    parser.add_argument("--dry-run", action="store_true",
                        help="Этап 1: только генерация JSON-сценариев (без сборки)")
    parser.add_argument("--top", type=int, default=10,
                        help="Количество видео для обработки (по умолч. 10)")
    parser.add_argument("--from-trend", type=str, default=None,
                        help="Путь к trend_top.json (если не указан — ищет в nura-trend-radar/data/)")
    parser.add_argument("--validate", action="store_true",
                        help="Проверить валидность всех сценариев в scenarios/")
    parser.add_argument("--assemble", action="store_true",
                        help="Этап 2: собрать видео по всем сценариям (после dry-run)")
    parser.add_argument("--to-xlsx", action="store_true",
                        help="Экспортировать все JSON-сценарии в Excel для просмотра")
    parser.add_argument("--parse-carousels", action="store_true",
                        help="Сгенерировать .carousel.json из carousel_texts/ (через AI parser)")
    parser.add_argument("--use-jobs", action="store_true",
                        help="Использовать новую job-структуру: jobs/{job_id}/input/media/work/output/")
    parser.add_argument("--from-job", type=str, default=None,
                        help="Собрать конкретный job по ID папки (jobs/{job_id})")
    parser.add_argument("--qa", action="store_true",
                        help="QA-проверка собранного видео/карусели")

    args = parser.parse_args()

    if args.validate:
        _validate_all()
        return

    if args.to_xlsx:
        _export_xlsx()
        return

    if args.parse_carousels:
        await _parse_carousels()
        return

    if args.from_job:
        await _assemble_single_job(args.from_job)
        return

    if args.qa:
        _run_qa_checks(args.from_job)
        return

    pipeline = VideoPipeline(
        trend_data_path=args.from_trend,
        dry_run=args.dry_run,
        use_jobs_dir=args.use_jobs,
    )

    if args.dry_run:
        print("=" * 60)
        mode = "jobs/" if args.use_jobs else "scenarios/"
        print(f"  NURA Pipeline — Этап 1: генерация ({mode})")
        print("=" * 60)
        generated = await pipeline.run(top_n=args.top)

        print(f"\nСгенерировано: {len(generated)}")
        for p in generated:
            if p.is_dir():
                valid = pipeline.validate_scenario(p / "input" / "scenario.json")
                print(f"  {'✓' if valid else '✗'} {p.name}")
            else:
                valid = pipeline.validate_scenario(p)
                print(f"  {'✓' if valid else '✗'} {p.name}")

        print("\nДалее:")
        print("  1. Запиши видео Нуры по текстам из nura_text")
        print("  2. Скачай стоковые видео по source_keywords в videos/media/")
        print("  3. Подправь JSON если нужно")
        print("  4. Запусти: python scripts/run_pipeline.py")
        print("     (или python scripts/assemble.py scenarios/имя.json для одного видео)")
    else:
        if args.assemble:
            print("=" * 60)
            print("  NURA Pipeline — Этап 2: сборка видео")
            print("=" * 60)
            if args.use_jobs:
                pipeline.generated = sorted(JOBS_DIR.glob("*/"))
            else:
                pipeline.generated = list((NURA_ROOT / "scenarios").glob("*.json"))
            pipeline.dry_run = False
            outputs = await pipeline.assemble_all()
            print(f"\nСобрано видео: {len(outputs)}")
            for o in outputs:
                print(f"  ✓ {o}")
        else:
            print("=" * 60)
            print("  NURA Pipeline — полный цикл")
            print("=" * 60)
            generated = await pipeline.run(top_n=args.top)
            outputs = await pipeline.assemble_all()
            print(f"\nГотово: {len(outputs)} видео")
            for o in outputs:
                print(f"  ✓ {o}")


def _validate_all():
    scenarios_dir = NURA_ROOT / "scenarios"
    from core.services.video_assembler import ScenarioConfig

    files = sorted(scenarios_dir.glob("*.json"))
    if not files:
        print("Нет файлов сценариев в scenarios/")
        return

    print(f"Валидация {len(files)} сценариев...")
    ok = 0
    for p in files:
        try:
            raw = json.loads(p.read_text("utf-8"))
            ScenarioConfig.model_validate(raw)
            print(f"  ✓ {p.name}")
            ok += 1
        except Exception as e:
            print(f"  ✗ {p.name}: {e}")
    print(f"\nВалидных: {ok}/{len(files)}")


def _export_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("Установи openpyxl: pip install openpyxl")
        return

    scenarios_dir = NURA_ROOT / "scenarios"
    files = sorted(scenarios_dir.glob("*.json"))
    if not files:
        print("Нет файлов сценариев в scenarios/")
        return

    out_dir = NURA_ROOT / "videos" / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "scenarios_view.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Сценарии NURA"

    headers = [
        ("Файл", 30), ("Сцена", 8), ("Старт", 10), ("Длит.", 10),
        ("Переход", 20), ("Текст озвучки (nura_text)", 70),
        ("Ключевые слова стока", 35), ("Оверлеи", 50),
        ("Цветокоррекция", 20),
    ]

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    cell_align = Alignment(vertical="top", wrap_text=True)

    for col, (name, w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[chr(64 + col)].width = w

    row = 2
    for fp in files:
        try:
            raw = json.loads(fp.read_text("utf-8"))
        except Exception:
            continue
        scenes = raw.get("scenes", [])
        if not scenes:
            ws.cell(row=row, column=1, value=fp.name)
            ws.cell(row=row, column=6, value="(нет сцен)")
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).alignment = cell_align
            row += 1
        for si, sc in enumerate(scenes, 1):
            ws.cell(row=row, column=1, value=fp.name)
            ws.cell(row=row, column=2, value=si)
            ws.cell(row=row, column=3, value=sc.get("start", 0))
            ws.cell(row=row, column=4, value=sc.get("duration", ""))
            tr = sc.get("transition")
            ws.cell(row=row, column=5, value=f"{tr['type']} / {tr['duration']}s" if tr else "")
            ws.cell(row=row, column=6, value=sc.get("nura_text", ""))
            ws.cell(row=row, column=7, value=", ".join(sc.get("source_keywords", [])))
            ovs = []
            for ov in sc.get("overlays", []):
                t = ov.get("type", "")
                if t == "text":
                    ovs.append(f"text: {ov.get('text', '')} [{ov.get('start')}-{ov.get('end')}s]")
                elif t == "zoom":
                    ovs.append(f"zoom: {ov.get('from_scale')}→{ov.get('to_scale')} ({ov.get('easing')})")
                elif t == "video":
                    ovs.append(f"video: {ov.get('src', '')}")
                elif t == "image":
                    ovs.append(f"image: {ov.get('src', '')}")
            ws.cell(row=row, column=8, value="\n".join(ovs))
            cg = sc.get("color_grading", {})
            ws.cell(row=row, column=9, value="вкл" if cg.get("enabled") else "—")
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).alignment = cell_align
            row += 1

    wb.save(str(out_path))
    print(f"Excel: {out_path}")


async def _parse_carousels():
    import httpx
    from pathlib import Path as _Path

    from core.config import settings

    prompts_dir = _Path(__file__).resolve().parent.parent / "nura_app" / "core" / "prompts"
    trend_data_dir = NURA_ROOT.parent / "nura-trend-radar" / "data" / "carousel_texts"
    if not trend_data_dir.exists():
        print("Нет файлов каруселей. Сначала запусти trend radar.")
        print(f"Ожидаемая папка: {trend_data_dir}")
        return

    files = sorted(trend_data_dir.glob("*.carousel.txt"))
    if not files:
        print("Нет .carousel.txt файлов для обработки.")
        return

    parser_prompt = (prompts_dir / "carousel_parser.txt").read_text("utf-8")
    scenarios_dir = NURA_ROOT / "scenarios"
    scenarios_dir.mkdir(parents=True, exist_ok=True)

    ok = 0
    for fp in files:
        name_base = fp.stem  # removes .carousel.txt
        carousel_json_path = scenarios_dir / f"{name_base}.carousel.json"
        if carousel_json_path.exists():
            print(f"  ✓ (уже есть) {carousel_json_path.name}")
            ok += 1
            continue

        text = fp.read_text("utf-8").strip()
        if not text:
            print(f"  ✗ пустой файл: {fp.name}")
            continue

        print(f"  → AI парсинг: {fp.name}...")
        try:
            async with httpx.AsyncClient(timeout=60.0) as client:
                r = await client.post(
                    f"{settings.deepseek_base_url}/chat/completions",
                    headers={
                        "Authorization": f"Bearer {settings.deepseek_api_key}",
                        "Content-Type": "application/json",
                    },
                    json={
                        "model": settings.deepseek_model,
                        "messages": [
                            {"role": "system", "content": parser_prompt},
                            {"role": "user", "content": text},
                        ],
                        "temperature": 0.3,
                        "max_tokens": 4000,
                    },
                )
                r.raise_for_status()
                raw = r.json()["choices"][0]["message"]["content"]

            parsed = _try_parse_json(raw)
            if parsed is None:
                print("    ✗ AI вернул не JSON")
                continue

            if parsed.get("error"):
                print(f"    ✗ AI: {parsed['error']}")
                continue

            with open(carousel_json_path, "w", encoding="utf-8") as f:
                json.dump(parsed, f, ensure_ascii=False, indent=2)
            print(f"    ✓ {carousel_json_path.name} ({len(parsed.get('slides', []))} слайдов)")
            ok += 1
        except Exception as e:
            print(f"    ✗ ошибка: {e}")

    print(f"\nГотово: {ok}/{len(files)} каруселей")


def _try_parse_json(text: str) -> dict | None:
    m = re.search(r"```(?:json)?\s*([\s\S]*?)```", text)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


async def _assemble_single_job(job_id: str):
    from core.services.video_assembler import ScenarioConfig, assemble
    from core.services.asset_validator import (
        validate_job_assets,
        format_asset_report,
    )
    from core.services.qa_checker import check_video_output, format_qa_result
    from core.services.packager import build_package

    job_dir = JOBS_DIR / job_id
    if not job_dir.exists():
        job_dir = Path(job_id)
        if not job_dir.exists():
            print(f"Job not found: {job_id}")
            return

    scenario_path = job_dir / "input" / "scenario.json"
    if not scenario_path.exists():
        print(f"No scenario.json in {job_dir}")
        return

    print("=" * 60)
    print(f"  Assembling job: {job_dir.name}")
    print("=" * 60)

    asset_report = validate_job_assets(job_dir, scenario_path)
    print(format_asset_report(asset_report))
    if not asset_report.passed:
        print("\n[FAIL] Fix asset issues and retry.")
        return

    raw = json.loads(scenario_path.read_text("utf-8"))
    cfg = ScenarioConfig.model_validate(raw)

    print("\n[1/4] Assembling video...")
    output_path = assemble(cfg, job_dir=job_dir)
    print(f"  Video: {output_path}")

    print("\n[2/4] QA check...")
    qa = check_video_output(output_path)
    print(format_qa_result(qa))

    print("\n[3/4] Packaging...")
    pkg = build_package(
        job_dir,
        video_path=output_path,
        carousel_dir=job_dir / "output" / "carousel",
    )
    print(f"  Package: {pkg.output_dir}")
    if pkg.video_path:
        print(f"  Video: {pkg.video_path.name}")
    if pkg.caption_path:
        print(f"  Caption: {pkg.caption_path.name}")
    if pkg.hashtags_path:
        print(f"  Hashtags: {pkg.hashtags_path.name}")
    if pkg.publish_notes_path:
        print(f"  Notes: {pkg.publish_notes_path.name}")

    print("\n[DONE]")


def _run_qa_checks(job_id: str | None = None):
    from core.services.qa_checker import (
        check_video_output,
        check_carousel_output,
        format_qa_result,
    )

    if job_id:
        job_dir = JOBS_DIR / job_id
        if not job_dir.exists():
            job_dir = Path(job_id)
            if not job_dir.exists():
                print(f"Job not found: {job_id}")
                return

        video_path = job_dir / "output" / f"{job_dir.name}.mp4"
        if not video_path.exists():
            mp4s = sorted(job_dir.glob("output/*.mp4"))
            video_path = mp4s[0] if mp4s else None

        if video_path and video_path.exists():
            print(format_qa_result(check_video_output(video_path)))
        else:
            print("No video found in output/")

        carousel_dir = job_dir / "output" / "carousel"
        if carousel_dir.exists():
            print()
            print(format_qa_result(check_carousel_output(carousel_dir), "Carousel"))
    else:
        videos_dir = NURA_ROOT / "videos" / "output"
        for mp4 in sorted(videos_dir.glob("*.mp4")):
            print(format_qa_result(check_video_output(mp4), mp4.name))
            print()


if __name__ == "__main__":
    asyncio.run(main())
